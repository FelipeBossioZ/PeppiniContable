# transactions/views/export.py
"""
Vista para exportación a Excel.
"""

from .base import *
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def export_to_excel_enhanced(request, company_id, year, month):
    """
    Exportación mejorada a Excel para partida doble.
    """
    try:
        company = Company.objects.get(id=company_id)
    except Company.DoesNotExist:
        return Response(
            {'error': 'Empresa no encontrada'},
            status=status.HTTP_404_NOT_FOUND
        )

    try:
        # Validar año y mes
        if not (2000 <= year <= 2100):
            raise ValueError("Año fuera de rango")
        if not (1 <= month <= 12):
            raise ValueError("Mes inválido")

        first_day_of_month = date(year, month, 1)
        last_day_of_month_num = monthrange(year, month)[1]
        last_day_of_month = date(year, month, last_day_of_month_num)

        # Obtener movimientos optimizados
        movements = Movement.objects.filter(
            transaction__company=company,
            transaction__date__lte=last_day_of_month
        ).select_related(
            'transaction', 'account', 'third_party'
        ).order_by('account__code', 'third_party__nit', 'transaction__date')

        # Calcular balances por cuenta y tercero
        balances = _calculate_balances(movements, first_day_of_month)

        # Crear libro de Excel
        workbook = _create_excel_workbook(
            company, year, month, first_day_of_month,
            last_day_of_month, balances
        )

        # Agregar hoja de detalle
        _add_detail_sheet(workbook, company, first_day_of_month, last_day_of_month)

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Libro_Diario_{company.nit}_{year}_{month:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        workbook.save(response)

        logger.info(f"Reporte Excel generado: {filename} por usuario {request.user.id}")

        return response

    except Exception as e:
        logger.error(f"Error generando reporte Excel: {str(e)}")
        return Response(
            {'error': f'Error al generar el reporte: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _calculate_balances(movements, first_day_of_month):
    """Calcula balances por cuenta y tercero."""
    balances = defaultdict(lambda: {
        'saldo_anterior': Decimal('0'),
        'debitos': Decimal('0'),
        'creditos': Decimal('0'),
        'movimientos': []
    })

    for m in movements:
        key = (m.account.code, m.account.name, m.third_party.nit, m.third_party.name)

        if m.transaction.date < first_day_of_month:
            balances[key]['saldo_anterior'] += m.debit - m.credit
        else:
            balances[key]['debitos'] += m.debit
            balances[key]['creditos'] += m.credit
            balances[key]['movimientos'].append({
                'fecha': m.transaction.date,
                'concepto': m.transaction.concept,
                'debito': m.debit,
                'credito': m.credit,
                'descripcion': m.description
            })

    return balances


def _create_excel_workbook(company, year, month, first_day, last_day, balances):
    """Crea el libro de Excel con el resumen."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Libro Diario"

    # Configurar estilos
    styles = _get_excel_styles()

    # Encabezados
    _add_excel_headers(worksheet, company, year, month, last_day, styles)

    # Encabezados de columnas
    headers = ["Cuenta", "Nombre Cuenta", "NIT Tercero", "Nombre Tercero",
               "Saldo Anterior", "Débitos", "Créditos", "Saldo Final"]

    row_num = 5
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col, value=header)
        cell.font = styles['column_header_font']
        cell.fill = styles['column_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = styles['thin_border']

    # Datos con formato
    row_num = 6
    totals = defaultdict(Decimal)

    for key, amounts in sorted(balances.items()):
        saldo_final = amounts['saldo_anterior'] + amounts['debitos'] - amounts['creditos']

        row_data = [
            key[0], key[1], key[2], key[3],
            float(amounts['saldo_anterior']),
            float(amounts['debitos']),
            float(amounts['creditos']),
            float(saldo_final)
        ]

        for col, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col, value=value)
            cell.border = styles['thin_border']

            if col >= 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                if value < 0:
                    cell.font = Font(color="E74C3C")
            else:
                cell.alignment = Alignment(horizontal='left')

            if row_num % 2 == 0:
                cell.fill = styles['alternating_fill']

        totals['saldo_anterior'] += amounts['saldo_anterior']
        totals['debitos'] += amounts['debitos']
        totals['creditos'] += amounts['creditos']
        totals['saldo_final'] += saldo_final

        row_num += 1

    # Línea de totales
    _add_totals_row(worksheet, row_num, totals, styles)

    # Ajustar anchos de columna
    column_widths = [15, 35, 15, 35, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width

    return workbook


def _get_excel_styles():
    """Retorna los estilos para el Excel."""
    return {
        'header_font': Font(bold=True, size=16, color="FFFFFF"),
        'subheader_font': Font(bold=True, size=12, color="FFFFFF"),
        'column_header_font': Font(bold=True, size=10, color="FFFFFF"),
        'header_fill': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
        'subheader_fill': PatternFill(start_color="34495E", end_color="34495E", fill_type="solid"),
        'column_fill': PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid"),
        'alternating_fill': PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid"),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def _add_excel_headers(worksheet, company, year, month, last_day, styles):
    """Agrega encabezados al Excel."""
    worksheet.merge_cells('A1:H1')
    cell = worksheet['A1']
    cell.value = f"{company.name} - NIT: {company.nit}"
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('A2:H2')
    cell = worksheet['A2']
    cell.value = "LIBRO DIARIO - PARTIDA DOBLE"
    cell.font = styles['subheader_font']
    cell.fill = styles['subheader_fill']
    cell.alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('A3:H3')
    cell = worksheet['A3']
    cell.value = f"Período: {month:02d}/{year} - Fecha de Corte: {last_day.strftime('%d/%m/%Y')}"
    cell.font = Font(italic=True, size=10)
    cell.alignment = Alignment(horizontal='center')

    worksheet.append([])


def _add_totals_row(worksheet, row_num, totals, styles):
    """Agrega la fila de totales."""
    row_num += 1
    worksheet.cell(row=row_num, column=3, value="TOTALES").font = Font(bold=True, size=12)

    for col, key in enumerate(['saldo_anterior', 'debitos', 'creditos', 'saldo_final'], 5):
        cell = worksheet.cell(row=row_num, column=col, value=float(totals[key]))
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
        cell.border = styles['thin_border']

    # Validación del balance
    balance_check = totals['debitos'] - totals['creditos']
    row_num += 2
    worksheet.cell(row=row_num, column=1, value="VALIDACIÓN DE BALANCE:").font = Font(bold=True)
    row_num += 1

    if abs(balance_check) < Decimal('0.01'):
        cell = worksheet.cell(row=row_num, column=1, value="✓ Balance cuadrado correctamente")
        cell.font = Font(color="27AE60", bold=True)
    else:
        cell = worksheet.cell(row=row_num, column=1, value=f"✗ Diferencia de balance: {float(balance_check):,.2f}")
        cell.font = Font(color="E74C3C", bold=True)


def _add_detail_sheet(workbook, company, first_day, last_day):
    """Agrega la hoja de detalle de movimientos del mes."""
    styles = _get_excel_styles()
    detail_sheet = workbook.create_sheet("Detalle del Mes")
    detail_headers = ["Fecha", "Comprobante", "Cuenta", "Tercero", "Concepto", "Descripción", "Débito", "Crédito"]

    for col, header in enumerate(detail_headers, 1):
        cell = detail_sheet.cell(row=1, column=col, value=header)
        cell.font = styles['column_header_font']
        cell.fill = styles['column_fill']
        cell.border = styles['thin_border']

    month_movements = Movement.objects.filter(
        transaction__company=company,
        transaction__date__range=[first_day, last_day]
    ).select_related('transaction', 'account', 'third_party').order_by('transaction__date')

    row_num = 2
    for m in month_movements:
        detail_sheet.append([
            m.transaction.date.strftime('%d/%m/%Y'),
            m.transaction.number,
            f"{m.account.code} - {m.account.name}",
            f"{m.third_party.nit} - {m.third_party.name}",
            m.transaction.concept,
            m.description or "",
            float(m.debit),
            float(m.credit)
        ])

        for col in range(1, len(detail_headers) + 1):
            cell = detail_sheet.cell(row=row_num, column=col)
            cell.border = styles['thin_border']
            if col >= 7:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if row_num % 2 == 0:
                cell.fill = styles['alternating_fill']

        row_num += 1

    detail_widths = [12, 15, 25, 25, 30, 30, 12, 12]
    for i, width in enumerate(detail_widths, 1):
        detail_sheet.column_dimensions[get_column_letter(i)].width = width


# Alias para compatibilidad
export_to_excel = export_to_excel_enhanced
