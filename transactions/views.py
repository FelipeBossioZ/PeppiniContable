# transactions/views.py 
from django.http import JsonResponse, HttpResponse
from django.db import transaction as db_transaction
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F, DecimalField
from django.utils import timezone
from .models import Transaction, Movement, Company, Account, ThirdParty, RecurringTransaction, AccountingRule
from .serializers import (
    TransactionSerializer, TransactionCreateSerializer, MovementSerializer,
    CompanySerializer, AccountSerializer, ThirdPartySerializer
)
from .permissions import HasValidLicense
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from collections import defaultdict
from decimal import Decimal
from datetime import date, timedelta
from calendar import monthrange
import logging

import pandas as pd
import zipfile
import rarfile
import re
from pathlib import Path

logger = logging.getLogger(__name__)

# ==============================================
# VISTAS MEJORADAS DE TRANSACCIONES CON PARTIDA DOBLE
# ==============================================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def transaction_list(request):
    """
    Lista y crea transacciones con múltiples movimientos (partida doble)
    """
    if request.method == 'GET':
        # Optimización con select_related para reducir queries
        queryset = Transaction.objects.select_related('company').prefetch_related(
            'movements__account', 'movements__third_party'
        ).order_by('-date', '-id')
        
        # Sistema de filtros avanzados
        filters = {}
        
        # Filtro por empresa
        if company_id := request.GET.get('company'):
            filters['company_id'] = company_id
        
        # Filtro por rango de fechas
        if start_date := request.GET.get('start_date'):
            filters['date__gte'] = start_date
        if end_date := request.GET.get('end_date'):
            filters['date__lte'] = end_date
        
        # Filtro por concepto
        if search := request.GET.get('search'):
            queryset = queryset.filter(
                Q(concept__icontains=search) | 
                Q(additional_description__icontains=search)
            )
        
        # Aplicar filtros
        if filters:
            queryset = queryset.filter(**filters)
        
        # Paginación mejorada
        page_size = min(int(request.GET.get('page_size', 50)), 100)
        paginator = Paginator(queryset, page_size)
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.page(page_number)
        except Exception:
            return Response(
                {'error': 'Página no válida'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Calcular totales para la página actual
        total_debits = 0
        total_credits = 0
        
        for transaction in page_obj:
            for movement in transaction.movements.all():
                total_debits += movement.debit
                total_credits += movement.credit
        
        serializer = TransactionSerializer(page_obj, many=True)
        
        return Response({
            'results': serializer.data,
            'pagination': {
                'count': paginator.count,
                'total_pages': paginator.num_pages,
                'current_page': page_obj.number,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'page_size': page_size
            },
            'totals': {
                'debits': float(total_debits),
                'credits': float(total_credits),
                'balance': float(total_debits - total_credits)
            }
        })

    elif request.method == 'POST':
        serializer = TransactionCreateSerializer(data=request.data)
    
        if serializer.is_valid():
            try:
                with db_transaction.atomic():
                    transaction_obj = serializer.save()
                    
                    # 🔥 NUEVO: Obtener alertas y enviarlas al frontend
                    alertas, sugerencias = transaction_obj.validar_logica_contable()
                    
                    # Invalidar caché
                    cache_keys = [f'dashboard_{transaction_obj.company_id}']
                    cache.delete_many(cache_keys)
                    
                    logger.info(f"Transacción creada: {transaction_obj.id}")
                    
                    # 🔥 NUEVO: Incluir alertas en la respuesta
                    response_data = TransactionSerializer(transaction_obj).data
                    if alertas:
                        response_data['alertas'] = alertas
                        response_data['sugerencias'] = sugerencias
                    
                    return Response(response_data, status=status.HTTP_201_CREATED)
                    
            except Exception as e:
                logger.error(f"Error creando transacción: {str(e)}")
                return Response(
                    {'error': f'Error al crear la transacción: {str(e)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ==============================================
# DASHBOARD CON ESTADÍSTICAS ACTUALIZADO
# ==============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def dashboard_stats(request):
    """
    Estadísticas del dashboard actualizadas para partida doble
    """
    company_id = request.GET.get('company')
    if not company_id:
        return Response(
            {'error': 'Se requiere el ID de la empresa'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Intentar obtener del caché
    cache_key = f'dashboard_{company_id}'
    cached_data = cache.get(cache_key)
    if cached_data:
        return Response(cached_data)
    
    try:
        today = date.today()
        first_day_month = date(today.year, today.month, 1)
        
        # Estadísticas del mes actual (desde Movements)
        current_month_movements = Movement.objects.filter(
            transaction__company_id=company_id,
            transaction__date__gte=first_day_month,
            transaction__date__lte=today
        )
        
        # Totales del mes
        month_totals = current_month_movements.aggregate(
            total_debits=Sum('debit'),
            total_credits=Sum('credit'),
            movement_count=Sum(1)
        )
        
        # Balance total histórico
        all_time_totals = Movement.objects.filter(
            transaction__company_id=company_id
        ).aggregate(
            total_debits=Sum('debit'),
            total_credits=Sum('credit')
        )
        
        # Top 5 cuentas más utilizadas
        top_accounts = (
            Movement.objects.filter(transaction__company_id=company_id)
            .values('account__name', 'account__code')
            .annotate(
                total=Sum(F('debit') + F('credit')),
                count=Sum(1)
            )
            .order_by('-total')[:5]
        )
        
        # Tendencia últimos 6 meses
        monthly_trend = []
        for i in range(5, -1, -1):
            month_date = today - timedelta(days=30*i)
            month_start = date(month_date.year, month_date.month, 1)
            month_end_day = monthrange(month_date.year, month_date.month)[1]
            month_end = date(month_date.year, month_date.month, month_end_day)
            
            month_data = Movement.objects.filter(
                transaction__company_id=company_id,
                transaction__date__range=[month_start, month_end]
            ).aggregate(
                debits=Sum('debit'),
                credits=Sum('credit')
            )
            
            monthly_trend.append({
                'month': month_start.strftime('%Y-%m'),
                'month_name': month_start.strftime('%B %Y'),
                'debits': float(month_data['debits'] or 0),
                'credits': float(month_data['credits'] or 0),
                'balance': float((month_data['debits'] or 0) - (month_data['credits'] or 0))
            })
        
        stats = {
            'current_month': {
                'debits': float(month_totals['total_debits'] or 0),
                'credits': float(month_totals['total_credits'] or 0),
                'movement_count': month_totals['movement_count'] or 0,
                'balance': float((month_totals['total_debits'] or 0) - (month_totals['total_credits'] or 0))
            },
            'all_time': {
                'debits': float(all_time_totals['total_debits'] or 0),
                'credits': float(all_time_totals['total_credits'] or 0),
                'balance': float((all_time_totals['total_debits'] or 0) - (all_time_totals['total_credits'] or 0))
            },
            'top_accounts': list(top_accounts),
            'monthly_trend': monthly_trend,
            'last_updated': timezone.now().isoformat()
        }
        
        # Guardar en caché por 5 minutos
        cache.set(cache_key, stats, 300)
        
        return Response(stats)
        
    except Exception as e:
        logger.error(f"Error generando estadísticas: {str(e)}")
        return Response(
            {'error': 'Error al generar estadísticas'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# ==============================================
# EXPORTACIÓN MEJORADA A EXCEL CON PARTIDA DOBLE
# ==============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def export_to_excel_enhanced(request, company_id, year, month):
    """
    Exportación mejorada a Excel para partida doble
    """
    try:
        company = Company.objects.get(id=company_id)
    except Company.DoesNotExist:
        return Response(
            {'error': 'Empresa no encontrada'}, 
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        # Validar año y mes
        if not (2000 <= year <= 2100):
            raise ValueError("Año fuera de rango")
        if not (1 <= month <= 12):
            raise ValueError("Mes inválido")
        
        first_day_of_month = date(year, month, 1)
        last_day_of_month_num = monthrange(year, month)[1]
        last_day_of_month = date(year, month, last_day_of_month_num)

        # Obtener movimientos optimizados
        movements = Movement.objects.filter(
            transaction__company=company,
            transaction__date__lte=last_day_of_month
        ).select_related(
            'transaction', 'account', 'third_party'
        ).order_by('account__code', 'third_party__nit', 'transaction__date')

        # Calcular balances por cuenta y tercero
        balances = defaultdict(lambda: {
            'saldo_anterior': Decimal('0'),
            'debitos': Decimal('0'),
            'creditos': Decimal('0'),
            'movimientos': []
        })
        
        for m in movements:
            key = (m.account.code, m.account.name, m.third_party.nit, m.third_party.name)
            
            if m.transaction.date < first_day_of_month:
                balances[key]['saldo_anterior'] += m.debit - m.credit
            else:
                balances[key]['debitos'] += m.debit
                balances[key]['creditos'] += m.credit
                balances[key]['movimientos'].append({
                    'fecha': m.transaction.date,
                    'concepto': m.transaction.concept,
                    'debito': m.debit,
                    'credito': m.credit,
                    'descripcion': m.description
                })

        # Crear libro de Excel con estilos profesionales
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Libro Diario"

        # Configurar estilos
        header_font = Font(bold=True, size=16, color="FFFFFF")
        subheader_font = Font(bold=True, size=12, color="FFFFFF")
        column_header_font = Font(bold=True, size=10, color="FFFFFF")
        
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        subheader_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        column_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        alternating_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Encabezados
        worksheet.merge_cells('A1:H1')
        cell = worksheet['A1']
        cell.value = f"{company.name} - NIT: {company.nit}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.merge_cells('A2:H2')
        cell = worksheet['A2']
        cell.value = "LIBRO DIARIO - PARTIDA DOBLE"
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.merge_cells('A3:H3')
        cell = worksheet['A3']
        cell.value = f"Período: {month:02d}/{year} - Fecha de Corte: {last_day_of_month.strftime('%d/%m/%Y')}"
        cell.font = Font(italic=True, size=10)
        cell.alignment = Alignment(horizontal='center')
        
        # Línea en blanco
        worksheet.append([])

        # Encabezados de columnas
        headers = ["Cuenta", "Nombre Cuenta", "NIT Tercero", "Nombre Tercero", 
                  "Saldo Anterior", "Débitos", "Créditos", "Saldo Final"]
        
        row_num = 5
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col, value=header)
            cell.font = column_header_font
            cell.fill = column_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos con formato
        row_num = 6
        totals = defaultdict(Decimal)
        
        for key, amounts in sorted(balances.items()):
            saldo_final = amounts['saldo_anterior'] + amounts['debitos'] - amounts['creditos']
            
            row_data = [
                key[0],  # Cuenta
                key[1],  # Nombre Cuenta
                key[2],  # NIT
                key[3],  # Nombre Tercero
                float(amounts['saldo_anterior']),
                float(amounts['debitos']),
                float(amounts['creditos']),
                float(saldo_final)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                
                # Formato numérico para columnas de montos
                if col >= 5:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    
                    # Colorear valores negativos
                    if value < 0:
                        cell.font = Font(color="E74C3C")
                else:
                    cell.alignment = Alignment(horizontal='left')
                
                # Filas alternadas
                if row_num % 2 == 0:
                    cell.fill = alternating_fill
            
            # Acumular totales
            totals['saldo_anterior'] += amounts['saldo_anterior']
            totals['debitos'] += amounts['debitos']
            totals['creditos'] += amounts['creditos']
            totals['saldo_final'] += saldo_final
            
            row_num += 1

        # Línea de totales
        row_num += 1
        worksheet.cell(row=row_num, column=3, value="TOTALES").font = Font(bold=True, size=12)
        
        for col, key in enumerate(['saldo_anterior', 'debitos', 'creditos', 'saldo_final'], 5):
            cell = worksheet.cell(row=row_num, column=col, value=float(totals[key]))
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00'
            cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            cell.border = thin_border

        # Ajustar anchos de columna
        column_widths = [15, 35, 15, 35, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Validación del balance
        balance_check = totals['debitos'] - totals['creditos']
        row_num += 2
        worksheet.cell(row=row_num, column=1, value="VALIDACIÓN DE BALANCE:").font = Font(bold=True)
        row_num += 1
        
        if abs(balance_check) < Decimal('0.01'):
            cell = worksheet.cell(row=row_num, column=1, value="✓ Balance cuadrado correctamente")
            cell.font = Font(color="27AE60", bold=True)
        else:
            cell = worksheet.cell(row=row_num, column=1, value=f"✗ Diferencia de balance: {float(balance_check):,.2f}")
            cell.font = Font(color="E74C3C", bold=True)

        # Crear segunda hoja con detalle de movimientos del mes
        detail_sheet = workbook.create_sheet("Detalle del Mes")
        detail_headers = ["Fecha", "Comprobante", "Cuenta", "Tercero", "Concepto", "Descripción", "Débito", "Crédito"]
        
        # Estilo para encabezados de detalle
        for col, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(row=1, column=col, value=header)
            cell.font = column_header_font
            cell.fill = column_fill
            cell.border = thin_border
        
        # Agregar movimientos del mes
        month_movements = Movement.objects.filter(
            transaction__company=company,
            transaction__date__range=[first_day_of_month, last_day_of_month]
        ).select_related('transaction', 'account', 'third_party').order_by('transaction__date')
        
        row_num = 2
        for m in month_movements:
            detail_sheet.append([
                m.transaction.date.strftime('%d/%m/%Y'),
                m.transaction.number,
                f"{m.account.code} - {m.account.name}",
                f"{m.third_party.nit} - {m.third_party.name}",
                m.transaction.concept,
                m.description or "",
                float(m.debit),
                float(m.credit)
            ])
            
            # Aplicar formato a la fila
            for col in range(1, len(detail_headers) + 1):
                cell = detail_sheet.cell(row=row_num, column=col)
                cell.border = thin_border
                if col >= 7:  # Columnas de montos
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                if row_num % 2 == 0:
                    cell.fill = alternating_fill
            
            row_num += 1

        # Ajustar anchos de columna en detalle
        detail_widths = [12, 15, 25, 25, 30, 30, 12, 12]
        for i, width in enumerate(detail_widths, 1):
            detail_sheet.column_dimensions[get_column_letter(i)].width = width

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Libro_Diario_{company.nit}_{year}_{month:02d}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        workbook.save(response)
        
        logger.info(f"Reporte Excel generado: {filename} por usuario {request.user.id}")
        
        return response
        
    except Exception as e:
        logger.error(f"Error generando reporte Excel: {str(e)}")
        return Response(
            {'error': f'Error al generar el reporte: {str(e)}'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

# ==============================================
# VISTAS BÁSICAS ACTUALIZADAS
# ==============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def company_list(request):
    """Lista de empresas"""
    companies = Company.objects.all()
    serializer = CompanySerializer(companies, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def account_list(request):
    """Lista de cuentas con jerarquía"""
    accounts = Account.objects.all().order_by('code')
    serializer = AccountSerializer(accounts, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def third_party_list(request):
    """Lista de terceros"""
    third_parties = ThirdParty.objects.all()
    serializer = ThirdPartySerializer(third_parties, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated, HasValidLicense])
def movement_list(request):
    """Lista de movimientos con filtros"""
    movements = Movement.objects.select_related(
        'transaction', 'account', 'third_party'
    ).order_by('-transaction__date', '-id')
    
    # Aplicar filtros
    if company_id := request.GET.get('company'):
        movements = movements.filter(transaction__company_id=company_id)
    
    if start_date := request.GET.get('start_date'):
        movements = movements.filter(transaction__date__gte=start_date)
    
    if end_date := request.GET.get('end_date'):
        movements = movements.filter(transaction__date__lte=end_date)
    
    if account_id := request.GET.get('account'):
        movements = movements.filter(account_id=account_id)
    
    serializer = MovementSerializer(movements, many=True)
    return Response(serializer.data)

# ==============================================
# TRANSACCIONES RECURRENTES (FALTANTE)
# ==============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def generate_recurring_transactions(request):
    """
    Genera transacciones recurrentes - PLACEHOLDER POR AHORA
    """
    return Response({
        'message': 'Funcionalidad de transacciones recurrentes en desarrollo',
        'created': 0,
        'skipped': 0,
        'errors': 0
    }, status=status.HTTP_200_OK)    


# ==============================================
# CORRECCIÓN AUTOMÁTICA DE TRANSACCIONES
# ==============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def corregir_transaccion(request, transaction_id):
    """
    🔥 CORRECCIÓN INTELIGENTE
    Solo corrige clases 3, 4 y 5 (Patrimonio, Ingresos, Gastos)
    """
    try:
        transaction = Transaction.objects.get(id=transaction_id)
        
        movimientos_corregidos = 0
        movimientos_detalle = []
        
        # Cuentas especiales
        CUENTAS_ESPECIALES_DEBITO = ['4175']
        CUENTAS_ESPECIALES_CREDITO = ['5905']
        
        for movimiento in transaction.movements.all():
            cuenta = movimiento.account
            tipo_cuenta = cuenta.tipo
            codigo_cuenta = cuenta.code
            
            movimiento_original = {
                'id': movimiento.id,
                'cuenta': f"{codigo_cuenta} - {cuenta.name}",
                'tipo': tipo_cuenta,
                'debito_original': float(movimiento.debit),
                'credito_original': float(movimiento.credit)
            }
            
            # 🔥 IGNORAR ACTIVOS Y PASIVOS
            if tipo_cuenta in ['ACTIVO', 'PASIVO']:
                movimientos_detalle.append({
                    **movimiento_original,
                    'corregido': False,
                    'razon': f"{tipo_cuenta} - No requiere corrección automática"
                })
                continue
            
            necesita_correccion = False
            razon = ""
            
            # Solo validar PATRIMONIO, INGRESOS y GASTOS
            if codigo_cuenta in CUENTAS_ESPECIALES_DEBITO:
                if movimiento.credit > 0 and tipo_cuenta == 'INGRESO':
                    movimiento.debit = movimiento.credit
                    movimiento.credit = Decimal('0')
                    necesita_correccion = True
                    razon = f"Cuenta especial {codigo_cuenta} (ingreso) debe aumentar con débito"
                    
            elif codigo_cuenta in CUENTAS_ESPECIALES_CREDITO:
                if movimiento.debit > 0 and tipo_cuenta == 'GASTO':
                    movimiento.credit = movimiento.debit
                    movimiento.debit = Decimal('0')
                    necesita_correccion = True
                    razon = f"Cuenta especial {codigo_cuenta} (gasto) debe aumentar con crédito"
                    
            else:
                # PATRIMONIO e INGRESOS: Aumentan con CRÉDITO
                if movimiento.debit > 0 and tipo_cuenta in ['PATRIMONIO', 'INGRESO']:
                    movimiento.credit = movimiento.debit
                    movimiento.debit = Decimal('0')
                    necesita_correccion = True
                    razon = f"{tipo_cuenta} normal debe aumentar con crédito"
                    
                # GASTOS: Aumentan con DÉBITO
                elif movimiento.credit > 0 and tipo_cuenta == 'GASTO':
                    movimiento.debit = movimiento.credit
                    movimiento.credit = Decimal('0')
                    necesita_correccion = True
                    razon = f"{tipo_cuenta} normal debe aumentar con débito"
            
            if necesita_correccion:
                movimiento.save()
                movimientos_corregidos += 1
            
            movimientos_detalle.append({
                **movimiento_original,
                'corregido': necesita_correccion,
                'razon': razon if necesita_correccion else 'Sin cambios'
            })
        
        # Validar balance
        total_debit = sum(m.debit for m in transaction.movements.all())
        total_credit = sum(m.credit for m in transaction.movements.all())
        diferencia = total_debit - total_credit
        
        return Response({
            'success': True,
            'message': f'Se corrigieron {movimientos_corregidos} movimientos',
            'movimientos_corregidos': movimientos_corregidos,
            'detalle_correcciones': movimientos_detalle,
            'balance_final': {
                'debitos': float(total_debit),
                'creditos': float(total_credit),
                'diferencia': float(diferencia),
                'balanceado': abs(diferencia) < Decimal('0.01')
            },
            'transaction': TransactionSerializer(transaction).data
        })
        
    except Transaction.DoesNotExist:
        return Response({'error': 'Transacción no encontrada'}, status=404)
    except Exception as e:
        logger.error(f"Error en corrección automática: {str(e)}")
        return Response({'error': f'Error interno: {str(e)}'}, status=500)

# ============================================
# AGREGAR ESTA FUNCIÓN AL FINAL DE views.py
# (antes del último comentario o al final del archivo)
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def calcular_correcciones(request, transaction_id):
    """
    Calcula cómo deberían estar los movimientos SIN guardar nada
    """
    try:
        transaction = Transaction.objects.get(id=transaction_id)
        
        correcciones = []
        
        # Cuentas especiales (casos raros)
        CUENTAS_ESPECIALES_DEBITO = ['4175']
        CUENTAS_ESPECIALES_CREDITO = ['5905']
        
        for movimiento in transaction.movements.all():
            cuenta = movimiento.account
            tipo_cuenta = cuenta.tipo
            codigo_cuenta = cuenta.code
            
            # Valores actuales (mal)
            debito_actual = float(movimiento.debit)
            credito_actual = float(movimiento.credit)
            
            # Valores corregidos (bien)
            debito_corregido = debito_actual
            credito_corregido = credito_actual
            
            # Lógica de corrección
            if codigo_cuenta in CUENTAS_ESPECIALES_DEBITO:
                if credito_actual > 0 and tipo_cuenta == 'INGRESO':
                    debito_corregido = credito_actual
                    credito_corregido = 0
                    
            elif codigo_cuenta in CUENTAS_ESPECIALES_CREDITO:
                if debito_actual > 0 and tipo_cuenta == 'GASTO':
                    credito_corregido = debito_actual
                    debito_corregido = 0
                    
            else:
                # Reglas normales
                if credito_actual > 0 and tipo_cuenta in ['ACTIVO', 'GASTO']:
                    debito_corregido = credito_actual
                    credito_corregido = 0
                    
                elif debito_actual > 0 and tipo_cuenta in ['PASIVO', 'INGRESO', 'PATRIMONIO']:
                    credito_corregido = debito_actual
                    debito_corregido = 0
            
            correcciones.append({
                'movement_index': list(transaction.movements.all()).index(movimiento),
                'account_id': cuenta.id,
                'third_party_id': movimiento.third_party.id,
                'debito_corregido': debito_corregido,
                'credito_corregido': credito_corregido
            })
        
        return Response({
            'correcciones': correcciones
        })
        
    except Transaction.DoesNotExist:
        return Response({'error': 'Transacción no encontrada'}, status=404)
    except Exception as e:
        logger.error(f"Error calculando correcciones: {str(e)}")
        return Response({'error': str(e)}, status=500)

# ============================================
# AGREGAR ESTA FUNCIÓN AL FINAL DE views.py
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def validate_transaction(request):
    """
    Valida un asiento SIN guardarlo en la base de datos
    Devuelve alertas y sugerencias si las hay
    
    🔥 SOLO VALIDA CLASES 3, 4 Y 5 (Patrimonio, Ingresos, Gastos)
    No valida Clase 1 (Activos) ni Clase 2 (Pasivos) porque es normal 
    que tengan débitos y créditos en asientos complejos
    """
    try:
        # Obtener los datos sin guardar
        movements_data = request.data.get('movements', [])
        
        alertas = []
        sugerencias = []
        correcciones = []
        
        # Cuentas especiales
        CUENTAS_ESPECIALES_DEBITO = ['4175']
        CUENTAS_ESPECIALES_CREDITO = ['5905']
        
        for index, mov_data in enumerate(movements_data):
            try:
                cuenta = Account.objects.get(id=mov_data['account'])
                tipo_cuenta = cuenta.tipo
                codigo_cuenta = cuenta.code
                
                debito = float(mov_data.get('debit', 0))
                credito = float(mov_data.get('credit', 0))
                
                # Valores corregidos
                debito_corregido = debito
                credito_corregido = credito
                
                # 🔥 NUEVO: Solo validar clases 3, 4 y 5
                # IGNORAR validaciones para ACTIVO (clase 1) y PASIVO (clase 2)
                if tipo_cuenta in ['ACTIVO', 'PASIVO']:
                    # No mostrar alertas para activos ni pasivos
                    correcciones.append({
                        'movement_index': index,
                        'account_id': cuenta.id,
                        'third_party_id': mov_data['third_party'],
                        'debito_corregido': debito,
                        'credito_corregido': credito
                    })
                    continue  # Saltar al siguiente movimiento
                
                # Validar SOLO para PATRIMONIO, INGRESOS y GASTOS
                if debito > 0:
                    # PATRIMONIO e INGRESOS normalmente aumentan con CRÉDITO
                    if tipo_cuenta in ['PATRIMONIO', 'INGRESO'] and codigo_cuenta not in CUENTAS_ESPECIALES_DEBITO:
                        alertas.append(f"⚠️ DÉBITO a {codigo_cuenta} - {cuenta.name} ({tipo_cuenta})")
                        sugerencias.append(f"Los {tipo_cuenta.lower()}s normalmente aumentan con CRÉDITO")
                        # Calcular corrección
                        credito_corregido = debito
                        debito_corregido = 0
                        
                elif credito > 0:
                    # GASTOS normalmente aumentan con DÉBITO
                    if tipo_cuenta == 'GASTO' and codigo_cuenta not in CUENTAS_ESPECIALES_CREDITO:
                        alertas.append(f"⚠️ CRÉDITO a {codigo_cuenta} - {cuenta.name} ({tipo_cuenta})")
                        sugerencias.append(f"Los gastos normalmente aumentan con DÉBITO")
                        # Calcular corrección
                        debito_corregido = credito
                        credito_corregido = 0
                
                correcciones.append({
                    'movement_index': index,
                    'account_id': cuenta.id,
                    'third_party_id': mov_data['third_party'],
                    'debito_corregido': debito_corregido,
                    'credito_corregido': credito_corregido
                })
                
            except Account.DoesNotExist:
                continue
        
        return Response({
            'alertas': alertas,
            'sugerencias': sugerencias,
            'correcciones': correcciones
        })
        
    except Exception as e:
        logger.error(f"Error validando transacción: {str(e)}")
        return Response({'error': str(e)}, status=500)

# ============================================
# 🆕 PROCESAMIENTO AUTOMÁTICO DE FACTURAS DIAN
# AGREGADO: [fecha de hoy]
# ============================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def procesar_facturas_dian_excel(request):
    """
    🎯 PROCESA ARCHIVOS EXCEL DE LA DIAN
    Este es el método RECOMENDADO - mucho más simple y confiable
    
    Archivos soportados:
    - Facturas_Recibidas_YYYYMM.xlsx (GASTOS)
    - Facturas_Emitidas_YYYYMM.xlsx (INGRESOS)
    """
    archivo = request.FILES.get('archivo')
    company_id = request.data.get('company')
    tipo = request.data.get('tipo')  # 'recibidas' o 'emitidas'
    
    if not archivo:
        return Response({'error': 'Debe subir un archivo'}, status=400)
    
    if not company_id:
        return Response({'error': 'Debe seleccionar una empresa'}, status=400)
    
    try:
        # Leer Excel
        df = pd.read_excel(archivo)
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.strip().str.lower()
        
        resultados = {
            'procesados': 0,
            'exitosos': 0,
            'errores': 0,
            'duplicados': 0,
            'detalles': []
        }
        
        # Procesar cada fila
        for index, row in df.iterrows():
            resultados['procesados'] += 1
            
            try:
                if tipo == 'recibidas':
                    # FACTURAS RECIBIDAS = GASTOS
                    resultado = crear_asiento_gasto_desde_dian(row, company_id)
                else:
                    # FACTURAS EMITIDAS = INGRESOS
                    resultado = crear_asiento_ingreso_desde_dian(row, company_id)
                
                if resultado['estado'] == 'exitoso':
                    resultados['exitosos'] += 1
                elif resultado['estado'] == 'duplicado':
                    resultados['duplicados'] += 1
                else:
                    resultados['errores'] += 1
                
                resultados['detalles'].append(resultado)
                
            except Exception as e:
                resultados['errores'] += 1
                resultados['detalles'].append({
                    'fila': index + 2,
                    'estado': 'error',
                    'mensaje': str(e)
                })
                logger.error(f"Error procesando fila {index}: {str(e)}")
        
        return Response(resultados)
        
    except Exception as e:
        logger.error(f"Error general procesando Excel: {str(e)}")
        return Response({'error': str(e)}, status=500)


def crear_asiento_gasto_desde_dian(row, company_id):
    """
    Crea asiento contable de GASTO desde factura recibida
    
    Columnas típicas del Excel de la DIAN (facturas recibidas):
    - NIT Proveedor / NIT Emisor
    - Razón Social Proveedor / Nombre Emisor
    - Número de Factura / Número Documento
    - Fecha
    - Valor Total / Total Factura
    - Concepto / Descripción
    """
    try:
        # Extraer datos (adaptado a diferentes formatos de DIAN)
        nit = str(row.get('nit proveedor') or row.get('nit emisor') or row.get('nit') or '').strip()
        nombre = str(row.get('razón social proveedor') or row.get('nombre emisor') or row.get('razón social') or row.get('razon social proveedor') or row.get('razon social') or '').strip()
        numero_factura = str(row.get('número de factura') or row.get('número documento') or row.get('numero de factura') or row.get('numero documento') or row.get('numero') or '').strip()
        
        # Intentar parsear fecha
        fecha_str = row.get('fecha') or row.get('fecha factura') or row.get('fecha emision') or row.get('fecha emisión')
        if pd.isna(fecha_str):
            fecha = date.today()
        else:
            fecha = pd.to_datetime(fecha_str)
            
        valor = float(row.get('valor total') or row.get('total factura') or row.get('total') or row.get('valor') or 0)
        concepto = str(row.get('concepto') or row.get('descripción') or row.get('descripcion') or row.get('observaciones') or 'Gasto general').strip()
        
        # Validar datos mínimos
        if not nit or not valor or valor <= 0:
            return {
                'factura': numero_factura or 'N/A',
                'estado': 'error',
                'mensaje': 'Datos incompletos o valor inválido'
            }
        
        # Verificar duplicado
        if verificar_factura_duplicada(numero_factura, nit, company_id):
            return {
                'factura': numero_factura,
                'estado': 'duplicado',
                'mensaje': 'Factura ya registrada'
            }
        
        # Obtener o crear tercero
        tercero = obtener_o_crear_tercero(nit, nombre)
        
        # Clasificar gasto automáticamente
        # 🤖 Clasificación inteligente con aprendizaje
        cuenta_gasto_code, es_anomalia, razon_clasificacion = clasificar_gasto_inteligente(
            nit, nombre, valor, company_id
        )
        cuenta_gasto = Account.objects.get(code=cuenta_gasto_code)
        cuenta_caja = Account.objects.get(code='1105')  # Caja
        
        # Crear transacción
        with db_transaction.atomic():
            transaction = Transaction.objects.create(
                company_id=company_id,
                date=fecha.date() if hasattr(fecha, 'date') else fecha,
                concept=f"Factura {numero_factura}: {concepto[:100]}",
                additional_description=f"Procesado automáticamente desde Excel DIAN - {nombre}"
            )
            
            # Movimiento 1: DÉBITO a cuenta de gasto
            Movement.objects.create(
                transaction=transaction,
                account=cuenta_gasto,
                third_party=tercero,
                debit=Decimal(str(valor)),
                credit=Decimal('0'),
                description=f"Factura {numero_factura}"
            )
            
            # Movimiento 2: CRÉDITO a caja
            Movement.objects.create(
                transaction=transaction,
                account=cuenta_caja,
                third_party=tercero,
                debit=Decimal('0'),
                credit=Decimal(str(valor)),
                description=f"Pago factura {numero_factura}"
            )
        
        return {
            'factura': numero_factura,
            'estado': 'exitoso',
            'transaccion_id': transaction.id,
            'tercero': nombre,
            'valor': valor,
            'cuenta': cuenta_gasto_code
        }
        
    except Exception as e:
        return {
            'factura': numero_factura if 'numero_factura' in locals() else 'N/A',
            'estado': 'error',
            'mensaje': str(e)
        }


def crear_asiento_ingreso_desde_dian(row, company_id):
    """
    Crea asiento contable de INGRESO desde factura emitida
    """
    try:
        # Extraer datos
        nit = str(row.get('nit adquiriente') or row.get('nit cliente') or row.get('nit') or '').strip()
        nombre = str(row.get('razón social adquiriente') or row.get('nombre cliente') or row.get('razón social') or row.get('razon social adquiriente') or row.get('razon social') or '').strip()
        numero_factura = str(row.get('número de factura') or row.get('número documento') or row.get('numero de factura') or row.get('numero') or '').strip()
        
        fecha_str = row.get('fecha') or row.get('fecha factura') or row.get('fecha emision')
        if pd.isna(fecha_str):
            fecha = date.today()
        else:
            fecha = pd.to_datetime(fecha_str)
            
        valor = float(row.get('valor total') or row.get('total factura') or row.get('total') or 0)
        concepto = str(row.get('concepto') or row.get('descripción') or row.get('descripcion') or 'Venta').strip()
        
        if not nit or not valor or valor <= 0:
            return {
                'factura': numero_factura or 'N/A',
                'estado': 'error',
                'mensaje': 'Datos incompletos'
            }
        
        # Verificar duplicado
        if verificar_factura_duplicada(numero_factura, nit, company_id):
            return {
                'factura': numero_factura,
                'estado': 'duplicado',
                'mensaje': 'Factura ya registrada'
            }
        
        # Obtener o crear tercero
        tercero = obtener_o_crear_tercero(nit, nombre)
        
        # Cuentas para ingreso
        cuenta_ingreso = Account.objects.get(code='4135')  # Comercio al por mayor y menor
        cuenta_caja = Account.objects.get(code='1105')  # Caja
        
        # Crear transacción
        with db_transaction.atomic():
            transaction = Transaction.objects.create(
                company_id=company_id,
                date=fecha.date() if hasattr(fecha, 'date') else fecha,
                concept=f"Factura venta {numero_factura}: {concepto[:100]}",
                additional_description=f"Procesado automáticamente desde Excel DIAN - {nombre}"
            )
            
            # Movimiento 1: DÉBITO a caja
            Movement.objects.create(
                transaction=transaction,
                account=cuenta_caja,
                third_party=tercero,
                debit=Decimal(str(valor)),
                credit=Decimal('0'),
                description=f"Cobro factura {numero_factura}"
            )
            
            # Movimiento 2: CRÉDITO a ingresos
            Movement.objects.create(
                transaction=transaction,
                account=cuenta_ingreso,
                third_party=tercero,
                debit=Decimal('0'),
                credit=Decimal(str(valor)),
                description=f"Venta factura {numero_factura}"
            )
        
        return {
            'factura': numero_factura,
            'estado': 'exitoso',
            'transaccion_id': transaction.id,
            'tercero': nombre,
            'valor': valor
        }
        
    except Exception as e:
        return {
            'factura': numero_factura if 'numero_factura' in locals() else 'N/A',
            'estado': 'error',
            'mensaje': str(e)
        }


def verificar_factura_duplicada(numero_factura, nit, company_id):
    """Verifica si la factura ya fue registrada"""
    if not numero_factura:
        return False
        
    existe = Transaction.objects.filter(
        company_id=company_id,
        concept__icontains=numero_factura
    ).exists()
    
    return existe


def obtener_o_crear_tercero(nit, nombre):
    """Crea tercero si no existe"""
    # Limpiar NIT
    nit_limpio = re.sub(r'[^\d]', '', str(nit))
    
    if not nit_limpio:
        nit_limpio = '000000000'
    
    if not nombre:
        nombre = f"Tercero {nit_limpio}"
    
    # Buscar o crear
    tercero, created = ThirdParty.objects.get_or_create(
        nit=nit_limpio,
        defaults={'name': nombre}
    )
    
    # Si existe pero el nombre cambió, actualizar
    if not created and tercero.name != nombre and nombre != f"Tercero {nit_limpio}":
        tercero.name = nombre
        tercero.save()
    
    return tercero


def clasificar_gasto_automatico(concepto):
    """
    Clasifica el gasto en una cuenta PUC basado en palabras clave
    """
    CLASIFICACION = {
        '5135': ['arriendo', 'alquiler', 'renta', 'arrendamiento', 'lease', 'canon'],
        '5120': ['honorarios', 'consultoría', 'asesoría', 'servicios profesionales', 'consultoria', 'asesoria'],
        '5195': ['publicidad', 'marketing', 'facebook', 'google', 'instagram', 'ads', 'pauta'],
        '5105': ['gasolina', 'combustible', 'peaje', 'transporte', 'uber', 'taxi', 'parqueadero'],
        '5115': ['celular', 'internet', 'telecomunicaciones', 'claro', 'movistar', 'tigo', 'telefono'],
        '5140': ['impuesto', 'gravamen', 'predial', 'vehicular', 'ica', 'reteica'],
        '5145': ['seguros', 'póliza', 'aseguradora', 'poliza'],
        '5160': ['mantenimiento', 'reparación', 'repuesto', 'reparacion'],
        '5205': ['papelería', 'oficina', 'útiles', 'elementos', 'papeleria', 'utiles'],
        '5110': ['salario', 'nomina', 'nómina', 'sueldo', 'prestaciones'],
    }
    
    concepto_lower = concepto.lower()
    
    for cuenta, keywords in CLASIFICACION.items():
        if any(keyword in concepto_lower for keyword in keywords):
            return cuenta
    
    # Por defecto: Gastos varios
    return '5195'


@api_view(['POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def procesar_archivo_comprimido(request):
    """
    🗜️ DESCOMPRIME Y PROCESA ZIP/RAR
    Útil si recibes facturas individuales comprimidas
    """
    archivo = request.FILES.get('archivo')
    company_id = request.data.get('company')
    
    if not archivo:
        return Response({'error': 'Debe subir un archivo'}, status=400)
    
    # Crear carpeta temporal
    temp_dir = Path('temp_facturas')
    temp_dir.mkdir(exist_ok=True)
    
    # Guardar archivo comprimido
    archivo_path = temp_dir / archivo.name
    with open(archivo_path, 'wb+') as destination:
        for chunk in archivo.chunks():
            destination.write(chunk)
    
    try:
        # Descomprimir
        if archivo.name.endswith('.zip'):
            with zipfile.ZipFile(archivo_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
        elif archivo.name.endswith('.rar'):
            with rarfile.RarFile(archivo_path, 'r') as rar_ref:
                rar_ref.extractall(temp_dir)
        else:
            return Response({'error': 'Formato no soportado. Use ZIP o RAR'}, status=400)
        
        resultados = {
            'procesados': 0,
            'archivos_encontrados': [],
            'mensaje': 'Archivos descomprimidos exitosamente'
        }
        
        # Listar archivos encontrados
        for file in temp_dir.rglob('*'):
            if file.is_file():
                resultados['archivos_encontrados'].append(str(file.name))
                resultados['procesados'] += 1
        
        # Limpiar archivos temporales
        import shutil
        shutil.rmtree(temp_dir)
        
        return Response(resultados)
        
    except Exception as e:
        logger.error(f"Error procesando archivo comprimido: {str(e)}")
        
        # Limpiar en caso de error
        import shutil
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
            
        return Response({'error': str(e)}, status=500)        

# ============================================
# 🤖 SISTEMA DE CLASIFICACIÓN INTELIGENTE
# ============================================

def clasificar_gasto_inteligente(nit, nombre, valor, company_id):
    """
    Clasificación inteligente de gastos usando reglas aprendidas
    
    Lógica:
    1. Busca si existe una regla para este NIT en esta empresa
    2. Si existe → valida si el monto es normal o anómalo
    3. Si no existe → usa clasificación por palabras clave
    4. Siempre fallback a 5195 DIVERSOS si hay problemas
    """
    try:
        from .models import AccountingRule, Account
        
        # 1. Buscar regla existente
        try:
            rule = AccountingRule.objects.get(
                company_id=company_id,
                third_party_nit=nit
            )
            
            # 2. Validar si el monto es anómalo
            if rule.is_amount_anomaly(Decimal(str(valor)), threshold=0.5):
                logger.warning(
                    f"⚠️ ANOMALÍA: NIT {nit} - Promedio: {rule.average_amount}, "
                    f"Actual: {valor} (>{50}% diferencia)"
                )
                # Monto anómalo → Mandar a revisión (5195)
                return '5195', True, f"Monto anómalo: promedio ${rule.average_amount:,.0f}"
            
            # 3. Monto normal → Usar la regla aprendida
            rule.update_statistics(Decimal(str(valor)))
            return rule.account.code, False, f"Regla aprendida (confianza: {rule.confidence_score})"
            
        except AccountingRule.DoesNotExist:
            # No existe regla → Clasificación por palabras clave
            cuenta_code = clasificar_por_palabras_clave(nombre)
            return cuenta_code, False, "Clasificación por palabras clave"
            
    except Exception as e:
        logger.error(f"Error en clasificación inteligente: {str(e)}")
        return '5195', False, "Error en clasificación"


def clasificar_por_palabras_clave(texto):
    """
    Clasificación tradicional por palabras clave
    Fallback cuando no hay reglas aprendidas
    """
    CLASIFICACION = {
        '5120': ['arriendo', 'alquiler', 'renta', 'arrendamiento', 'lease', 'canon'],
        '5105': ['honorarios', 'nomina', 'nómina', 'salario', 'sueldo', 'prestaciones', 'personal'],
        '5135': ['servicios', 'mantenimiento', 'reparación', 'reparacion', 'aseo', 'vigilancia'],
        '5140': ['impuesto', 'gravamen', 'predial', 'vehicular', 'ica', 'reteica', 'iva'],
        '5145': ['seguros', 'póliza', 'aseguradora', 'poliza', 'seguro'],
        '5115': ['celular', 'internet', 'telecomunicaciones', 'telefono', 'datos'],
    }
    
    texto_lower = texto.lower()
    
    for cuenta, keywords in CLASIFICACION.items():
        if any(keyword in texto_lower for keyword in keywords):
            return cuenta
    
    # Default: Gastos diversos
    return '5195'


def aprender_de_edicion(transaction_id, movement_id, nueva_cuenta_id):
    """
    Aprende automáticamente cuando el usuario edita una transacción
    Crea o actualiza reglas en AccountingRule
    """
    try:
        from .models import Movement, AccountingRule, Account
        
        movement = Movement.objects.select_related(
            'transaction', 'third_party', 'account'
        ).get(id=movement_id)
        
        nit = movement.third_party.nit
        nombre = movement.third_party.name
        company_id = movement.transaction.company_id
        valor = movement.debit if movement.debit > 0 else movement.credit
        nueva_cuenta = Account.objects.get(id=nueva_cuenta_id)
        
        # Crear o actualizar regla
        rule, created = AccountingRule.objects.get_or_create(
            company_id=company_id,
            third_party_nit=nit,
            defaults={
                'third_party_name': nombre,
                'account': nueva_cuenta,
                'created_by_user': True,
                'last_amount': valor,
                'average_amount': valor,
                'confidence_score': 1
            }
        )
        
        if not created:
            # Actualizar regla existente
            rule.account = nueva_cuenta
            rule.update_statistics(valor)
            rule.save()
            
            logger.info(
                f"✅ Regla actualizada: NIT {nit} → {nueva_cuenta.code} "
                f"(confianza: {rule.confidence_score})"
            )
        else:
            logger.info(f"🆕 Nueva regla creada: NIT {nit} → {nueva_cuenta.code}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error aprendiendo de edición: {str(e)}")
        return False


# ============================================
# 🔧 EDITAR TRANSACCIONES
# ============================================

@api_view(['PUT'])
@permission_classes([IsAuthenticated, HasValidLicense])
def edit_movement(request, movement_id):
    """
    Edita un movimiento individual y aprende de los cambios
    """
    try:
        movement = Movement.objects.select_related('transaction').get(id=movement_id)
        
        # Datos anteriores para el aprendizaje
        old_account_id = movement.account_id
        
        # Actualizar campos
        if 'account' in request.data:
            movement.account_id = request.data['account']
        if 'third_party' in request.data:
            movement.third_party_id = request.data['third_party']
        if 'debit' in request.data:
            movement.debit = Decimal(str(request.data['debit']))
        if 'credit' in request.data:
            movement.credit = Decimal(str(request.data['credit']))
        if 'description' in request.data:
            movement.description = request.data['description']
        
        movement.save()
        
        # 🤖 APRENDIZAJE AUTOMÁTICO
        if 'account' in request.data and old_account_id != movement.account_id:
            aprender_de_edicion(
                movement.transaction_id,
                movement_id,
                movement.account_id
            )
        
        # Invalidar caché
        cache.delete(f'dashboard_{movement.transaction.company_id}')
        
        return Response({
            'success': True,
            'message': 'Movimiento actualizado exitosamente',
            'movement': MovementSerializer(movement).data
        })
        
    except Movement.DoesNotExist:
        return Response({'error': 'Movimiento no encontrado'}, status=404)
    except Exception as e:
        logger.error(f"Error editando movimiento: {str(e)}")
        return Response({'error': str(e)}, status=500)


@api_view(['PUT'])
@permission_classes([IsAuthenticated, HasValidLicense])
def edit_transaction(request, transaction_id):
    """
    Edita los datos generales de una transacción
    """
    try:
        transaction = Transaction.objects.get(id=transaction_id)
        
        if 'date' in request.data:
            transaction.date = request.data['date']
        if 'concept' in request.data:
            transaction.concept = request.data['concept']
        if 'additional_description' in request.data:
            transaction.additional_description = request.data['additional_description']
        
        transaction.save()
        
        cache.delete(f'dashboard_{transaction.company_id}')
        
        return Response({
            'success': True,
            'message': 'Transacción actualizada exitosamente',
            'transaction': TransactionSerializer(transaction).data
        })
        
    except Transaction.DoesNotExist:
        return Response({'error': 'Transacción no encontrada'}, status=404)
    except Exception as e:
        logger.error(f"Error editando transacción: {str(e)}")
        return Response({'error': str(e)}, status=500)


# ============================================
# 📊 GESTIÓN DE REGLAS MANUALES
# ============================================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, HasValidLicense])
def accounting_rules_list(request):
    """
    Lista y crea reglas de clasificación manualmente
    """
    if request.method == 'GET':
        company_id = request.GET.get('company')
        if not company_id:
            return Response({'error': 'Se requiere company_id'}, status=400)
        
        rules = AccountingRule.objects.filter(
            company_id=company_id
        ).select_related('account').order_by('-confidence_score', 'third_party_name')
        
        rules_data = [{
            'id': rule.id,
            'third_party_nit': rule.third_party_nit,
            'third_party_name': rule.third_party_name,
            'account_code': rule.account.code,
            'account_name': rule.account.name,
            'confidence_score': rule.confidence_score,
            'average_amount': float(rule.average_amount) if rule.average_amount else None,
            'created_by_user': rule.created_by_user,
            'updated_at': rule.updated_at.isoformat()
        } for rule in rules]
        
        return Response(rules_data)
    
    elif request.method == 'POST':
               
        try:
            rule = AccountingRule.objects.create(
                company_id=request.data['company'],
                third_party_nit=request.data['nit'],
                third_party_name=request.data.get('name', ''),
                account_id=request.data['account'],
                created_by_user=True
            )
            
            return Response({
                'success': True,
                'message': 'Regla creada exitosamente',
                'rule_id': rule.id
            }, status=201)
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, HasValidLicense])
def delete_accounting_rule(request, rule_id):
    """
    Elimina una regla de clasificación
    """
    try:
        rule = AccountingRule.objects.get(id=rule_id)
        rule.delete()
        
        return Response({
            'success': True,
            'message': 'Regla eliminada exitosamente'
        })
        
    except AccountingRule.DoesNotExist:
        return Response({'error': 'Regla no encontrada'}, status=404)

# Alias para compatibilidad
export_to_excel = export_to_excel_enhanced